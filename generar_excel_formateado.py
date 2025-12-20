#!/usr/bin/env python3
"""
Generador de Excel Formateado para Comparativas
Crea archivos Excel con formato visual en las carpetas de comparativas
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def generar_excel_formateado(archivo_csv, carpeta_salida, nombre_archivo):
    """
    Genera Excel formateado exactamente como el original
    Formato: Materia/Persona | Lunes | Martes | Miércoles | Jueves | Viernes | Sábado
    
    Args:
        archivo_csv: Ruta al archivo CSV estructurado
        carpeta_salida: Carpeta donde guardar el Excel
        nombre_archivo: Nombre del archivo de salida (sin extensión)
    """
    print(f"\n📊 Generando Excel formateado: {nombre_archivo}")
    
    # Leer CSV estructurado
    df = pd.read_csv(archivo_csv, encoding='utf-8')
    
    # Agrupar por Grupo y Materia para reconstruir formato original
    grupos_materias = df.groupby(['Grupo', 'Materia', 'Profesor'])
    
    # Crear estructura para Excel
    datos_excel = []
    
    # Ordenar por grupo
    grupos_ordenados = sorted(grupos_materias.groups.keys(), key=lambda x: x[0])
    
    for (grupo, materia, profesor) in grupos_ordenados:
        grupo_df = grupos_materias.get_group((grupo, materia, profesor))
        
        # Crear fila combinando grupo, profesor y materia
        # Formato: "GRUPO\nPROFESOR\nMATERIA"
        materia_persona = f"{grupo}\n{profesor}\n{materia}"
        
        fila = {
            'Materia/Persona': materia_persona,
            'Lunes': '',
            'Martes': '',
            'Miercoles': '',
            'Jueves': '',
            'Viernes': '',
            'Sabado': ''
        }
        
        # Llenar horarios por día
        for _, row in grupo_df.iterrows():
            dia = row['Dia']
            bloque = str(row['Bloque_Horario']).zfill(4)  # Asegurar 4 dígitos
            salon = row['Salon']
            
            # Formatear hora: "0708" -> "7:00AM - 8:00AM"
            if len(bloque) == 4:
                hora_inicio_num = int(bloque[:2])
                hora_fin_num = int(bloque[2:])
                
                # Convertir a formato 12 horas con AM/PM
                if hora_inicio_num == 0:
                    hora_inicio_str = "12:00AM"
                elif hora_inicio_num < 12:
                    hora_inicio_str = f"{hora_inicio_num}:00AM"
                elif hora_inicio_num == 12:
                    hora_inicio_str = "12:00PM"
                else:
                    hora_inicio_str = f"{hora_inicio_num - 12}:00PM"
                
                if hora_fin_num == 0:
                    hora_fin_str = "12:00AM"
                elif hora_fin_num < 12:
                    hora_fin_str = f"{hora_fin_num}:00AM"
                elif hora_fin_num == 12:
                    hora_fin_str = "12:00PM"
                else:
                    hora_fin_str = f"{hora_fin_num - 12}:00PM"
                
                hora_formateada = f"{hora_inicio_str} - {hora_fin_str}"
            else:
                hora_formateada = bloque
            
            # Crear celda: "7:00AM - 8:00AM SALON"
            celda = f"{hora_formateada} {salon}"
            
            # Asignar a día correspondiente
            if dia in fila:
                if fila[dia]:  # Si ya hay algo, agregar en nueva línea
                    fila[dia] += f" {celda}"
                else:
                    fila[dia] = celda
        
        datos_excel.append(fila)
    
    # Crear DataFrame para Excel
    df_excel = pd.DataFrame(datos_excel)
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"
    
    # Escribir datos
    for r_idx, row in enumerate(dataframe_to_rows(df_excel, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Estilos
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Colores alternados para filas (gris claro)
    row_fill_light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    invalido_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    invalido_font = Font(color="FFFFFF", bold=True)
    
    border_thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Aplicar formato a encabezado (negro con texto blanco)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    
    # Aplicar formato a datos
    invalidos_count = 0
    for row_idx in range(2, ws.max_row + 1):
        # Alternar colores de fila
        if row_idx % 2 == 0:
            row_fill = row_fill_light
        else:
            row_fill = row_fill_white
        
        tiene_invalido = False
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = str(cell.value) if cell.value else ""
            
            # Detectar salones inválidos
            if any(inv in cell_value for inv in ['AV1', 'AV2', 'AV4', 'AV5', 'E11']):
                tiene_invalido = True
                invalidos_count += cell_value.count('AM') + cell_value.count('PM')  # Contar asignaciones
                cell.fill = invalido_fill
                cell.font = invalido_font
            else:
                cell.fill = row_fill
            
            # Alineación
            if col_idx == 1:  # Primera columna (Materia/Persona)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cell.border = border_thin
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 40  # Materia/Persona (más ancho)
    ws.column_dimensions['B'].width = 25  # Lunes
    ws.column_dimensions['C'].width = 25  # Martes
    ws.column_dimensions['D'].width = 25  # Miércoles
    ws.column_dimensions['E'].width = 25  # Jueves
    ws.column_dimensions['F'].width = 25  # Viernes
    ws.column_dimensions['G'].width = 15  # Sábado
    
    # Ajustar altura de filas (más alto para wrap text)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 45
    
    # Altura del encabezado
    ws.row_dimensions[1].height = 25
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar
    os.makedirs(carpeta_salida, exist_ok=True)
    output_path = os.path.join(carpeta_salida, f"{nombre_archivo}.xlsx")
    wb.save(output_path)
    
    print(f"   ✅ Excel guardado: {output_path}")
    print(f"   📏 Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
    print(f"   ⚠️  Asignaciones inválidas: {invalidos_count}")
    
    return output_path
