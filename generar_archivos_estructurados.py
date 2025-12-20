#!/usr/bin/env python3
"""
Generador de Archivos Estructurados (Excel y CSV)
Convierte los horarios a formato estructurado para análisis
"""

import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def parsear_horario(horario_str):
    """Parsea '0809/FF2' -> ('0809', 'FF2')"""
    if pd.isna(horario_str) or horario_str == '':
        return None, None
    match = re.match(r'(\d{4})/([A-Z0-9]+)', str(horario_str))
    if match:
        return match.group(1), match.group(2)
    return None, None

def convertir_a_estructura_plana(archivo_csv, nombre_archivo):
    """Convierte el CSV a estructura plana para análisis"""
    df = pd.read_csv(archivo_csv, encoding='utf-8')
    
    # Detectar nombre de columna
    materia_col = 'Materia.' if 'Materia.' in df.columns else 'Materia'
    
    registros = []
    profesor_actual = None
    
    for idx, row in df.iterrows():
        grupo = str(row['Grupo'])
        
        # Detectar profesor
        if 'PROFESOR' in grupo:
            profesor_actual = grupo.strip()
            continue
        
        # Saltar líneas vacías
        if pd.isna(row[materia_col]) or grupo == 'nan':
            continue
        
        materia = str(row[materia_col]).strip()
        horas_semana = row['Class'] if 'Class' in row and not pd.isna(row['Class']) else 0
        
        # Procesar cada día
        for dia in ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']:
            hora, salon = parsear_horario(row[dia])
            
            if hora and salon:
                # Determinar tipo de salón
                if salon.startswith('L'):
                    tipo_salon = 'Laboratorio'
                    piso = 'Primer Piso' if salon in {'LR', 'LSO', 'LIA', 'LCG1', 'LCG2'} else 'Segundo Piso'
                elif salon in {'AV1', 'AV2', 'AV4', 'AV5', 'E11'}:
                    tipo_salon = 'INVÁLIDO'
                    piso = 'N/A'
                else:
                    tipo_salon = 'Teoría'
                    piso = 'Planta Baja' if salon in {'FF1', 'FF2', 'FF3', 'FF4', 'FF5', 'FF6', 'FF7'} else 'Planta Alta'
                
                # Extraer hora inicio y fin
                hora_inicio = f"{hora[:2]}:{hora[2:]}"
                hora_fin_num = int(hora[2:]) + 1
                hora_fin = f"{hora[:2]}:{hora_fin_num:02d}"
                
                registro = {
                    'Archivo': nombre_archivo,
                    'Profesor': profesor_actual if profesor_actual else 'SIN ASIGNAR',
                    'Grupo': grupo,
                    'Materia': materia,
                    'Horas_Semana': int(horas_semana),
                    'Dia': dia,
                    'Hora_Inicio': hora_inicio,
                    'Hora_Fin': hora_fin,
                    'Bloque_Horario': hora,
                    'Salon': salon,
                    'Tipo_Salon': tipo_salon,
                    'Piso': piso,
                    'Es_Invalido': 1 if tipo_salon == 'INVÁLIDO' else 0,
                    'Es_Primer_Semestre': 1 if grupo.startswith('1') else 0
                }
                
                registros.append(registro)
    
    return pd.DataFrame(registros)

def crear_hoja_resumen(df, nombre):
    """Crea DataFrame de resumen estadístico"""
    resumen = {
        'Métrica': [],
        'Valor': []
    }
    
    resumen['Métrica'].append('Total de Asignaciones')
    resumen['Valor'].append(len(df))
    
    resumen['Métrica'].append('Materias Únicas')
    resumen['Valor'].append(df['Materia'].nunique())
    
    resumen['Métrica'].append('Grupos Únicos')
    resumen['Valor'].append(df['Grupo'].nunique())
    
    resumen['Métrica'].append('Profesores')
    resumen['Valor'].append(df['Profesor'].nunique())
    
    resumen['Métrica'].append('Salones Utilizados')
    resumen['Valor'].append(df['Salon'].nunique())
    
    resumen['Métrica'].append('⚠️ Asignaciones Inválidas')
    resumen['Valor'].append(df['Es_Invalido'].sum())
    
    resumen['Métrica'].append('Salones Teoría Usados')
    resumen['Valor'].append(df[df['Tipo_Salon'] == 'Teoría']['Salon'].nunique())
    
    resumen['Métrica'].append('Laboratorios Usados')
    resumen['Valor'].append(df[df['Tipo_Salon'] == 'Laboratorio']['Salon'].nunique())
    
    resumen['Métrica'].append('Grupos 1er Semestre')
    resumen['Valor'].append(df[df['Es_Primer_Semestre'] == 1]['Grupo'].nunique())
    
    return pd.DataFrame(resumen)

def crear_hoja_salones_invalidos(df):
    """Crea DataFrame con detalle de salones inválidos"""
    invalidos = df[df['Es_Invalido'] == 1].copy()
    
    if len(invalidos) == 0:
        return pd.DataFrame({'Mensaje': ['✅ No hay asignaciones inválidas']})
    
    return invalidos[['Grupo', 'Materia', 'Dia', 'Hora_Inicio', 'Salon', 'Profesor']].sort_values(['Salon', 'Grupo'])

def crear_hoja_por_materia(df):
    """Agrupa por materia"""
    por_materia = df.groupby('Materia').agg({
        'Grupo': 'nunique',
        'Salon': lambda x: ', '.join(sorted(set(x))),
        'Tipo_Salon': lambda x: ', '.join(sorted(set(x))),
        'Es_Invalido': 'sum',
        'Horas_Semana': 'first'
    }).reset_index()
    
    por_materia.columns = ['Materia', 'Num_Grupos', 'Salones_Usados', 'Tipos_Salon', 'Asig_Invalidas', 'Horas_Semana']
    
    return por_materia.sort_values('Num_Grupos', ascending=False)

def crear_hoja_por_salon(df):
    """Agrupa por salón"""
    por_salon = df.groupby('Salon').agg({
        'Grupo': 'nunique',
        'Materia': 'nunique',
        'Profesor': 'nunique',
        'Tipo_Salon': 'first',
        'Piso': 'first'
    }).reset_index()
    
    por_salon.columns = ['Salon', 'Num_Grupos', 'Num_Materias', 'Num_Profesores', 'Tipo', 'Piso']
    
    return por_salon.sort_values('Num_Grupos', ascending=False)

def aplicar_estilo_excel(wb, nombre_hoja, df, es_resumen=False):
    """Aplica estilos a una hoja de Excel"""
    ws = wb[nombre_hoja]
    
    # Estilo de encabezado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Aplicar estilo a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar bordes a todas las celdas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Resaltar valores inválidos
    if 'Es_Invalido' in df.columns or 'Asig_Invalidas' in df.columns:
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                col_name = ws.cell(1, col_idx).value
                if col_name in ['Es_Invalido', 'Asig_Invalidas'] and cell.value and int(cell.value) > 0:
                    cell.fill = red_fill
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'

def generar_archivos_estructurados(archivo_csv, nombre_salida):
    """Genera archivos Excel y CSV estructurados"""
    print(f"\n📊 Procesando: {archivo_csv}")
    
    # Convertir a estructura plana
    df_plano = convertir_a_estructura_plana(archivo_csv, nombre_salida)
    
    print(f"   ✅ {len(df_plano)} registros procesados")
    
    # Crear Excel con múltiples hojas
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto
    
    # Hoja 1: Resumen
    df_resumen = crear_hoja_resumen(df_plano, nombre_salida)
    ws_resumen = wb.create_sheet("Resumen")
    for r in dataframe_to_rows(df_resumen, index=False, header=True):
        ws_resumen.append(r)
    aplicar_estilo_excel(wb, "Resumen", df_resumen, es_resumen=True)
    
    # Hoja 2: Datos Completos
    ws_datos = wb.create_sheet("Datos_Completos")
    for r in dataframe_to_rows(df_plano, index=False, header=True):
        ws_datos.append(r)
    aplicar_estilo_excel(wb, "Datos_Completos", df_plano)
    
    # Hoja 3: Salones Inválidos
    df_invalidos = crear_hoja_salones_invalidos(df_plano)
    ws_invalidos = wb.create_sheet("Salones_Invalidos")
    for r in dataframe_to_rows(df_invalidos, index=False, header=True):
        ws_invalidos.append(r)
    aplicar_estilo_excel(wb, "Salones_Invalidos", df_invalidos)
    
    # Hoja 4: Por Materia
    df_materia = crear_hoja_por_materia(df_plano)
    ws_materia = wb.create_sheet("Por_Materia")
    for r in dataframe_to_rows(df_materia, index=False, header=True):
        ws_materia.append(r)
    aplicar_estilo_excel(wb, "Por_Materia", df_materia)
    
    # Hoja 5: Por Salón
    df_salon = crear_hoja_por_salon(df_plano)
    ws_salon = wb.create_sheet("Por_Salon")
    for r in dataframe_to_rows(df_salon, index=False, header=True):
        ws_salon.append(r)
    aplicar_estilo_excel(wb, "Por_Salon", df_salon)
    
    # Guardar Excel
    excel_path = f"/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/Sistema-Salones-ISC/datos_estructurados/{nombre_salida}.xlsx"
    wb.save(excel_path)
    print(f"   ✅ Excel guardado: {excel_path}")
    
    # Guardar CSV
    csv_path = f"/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/Sistema-Salones-ISC/datos_estructurados/{nombre_salida}.csv"
    df_plano.to_csv(csv_path, index=False, encoding='utf-8')
    print(f"   ✅ CSV guardado: {csv_path}")
    
    return df_plano

def main():
    """Función principal"""
    print("🚀 GENERADOR DE ARCHIVOS ESTRUCTURADOS")
    print("="*80)
    
    # Crear directorio de salida
    import os
    os.makedirs("/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/Sistema-Salones-ISC/datos_estructurados", exist_ok=True)
    
    # Procesar archivo original
    archivo_original = "/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/PROBLEMA SALONES ISC TEC/HorariosAgoDic2025.csv"
    df_original = generar_archivos_estructurados(archivo_original, "01_Horario_Inicial")
    
    # Procesar archivo optimizado
    archivo_optimizado = "/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/PROBLEMA SALONES ISC TEC/grouped_optimized_schedule.csv"
    df_optimizado = generar_archivos_estructurados(archivo_optimizado, "02_Horario_Optimizado_Profesor")
    
    print("\n" + "="*80)
    print("✅ ARCHIVOS GENERADOS EXITOSAMENTE")
    print("="*80)
    print("\n📁 Ubicación: Sistema-Salones-ISC/datos_estructurados/")
    print("\n📊 Archivos creados:")
    print("   1. 01_Horario_Inicial.xlsx (5 hojas)")
    print("   2. 01_Horario_Inicial.csv")
    print("   3. 02_Horario_Optimizado_Profesor.xlsx (5 hojas)")
    print("   4. 02_Horario_Optimizado_Profesor.csv")
    
    print("\n📋 Hojas de Excel:")
    print("   • Resumen - Estadísticas generales")
    print("   • Datos_Completos - Todos los registros")
    print("   • Salones_Invalidos - Asignaciones a AV/E11")
    print("   • Por_Materia - Agrupado por materia")
    print("   • Por_Salon - Agrupado por salón")
    
    print("\n🎯 Próximo paso: Análisis comparativo y gráficos\n")

if __name__ == "__main__":
    main()
