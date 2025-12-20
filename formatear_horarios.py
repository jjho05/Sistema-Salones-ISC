#!/usr/bin/env python3
"""
Formateador de Horarios - Convierte CSV a Excel con Formato Visual
Mantiene la estructura original pero mejora la legibilidad
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def aplicar_formato_horario(archivo_csv, nombre_salida):
    """Convierte CSV a Excel con formato mejorado"""
    print(f"\n📊 Formateando: {archivo_csv}")
    
    # Leer CSV
    df = pd.read_csv(archivo_csv, encoding='utf-8')
    
    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"
    
    # Escribir datos
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    profesor_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    profesor_font = Font(bold=True, size=11)
    invalido_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    invalido_font = Font(color="FFFFFF", bold=True)
    
    border_thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Aplicar formato a encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    
    # Aplicar formato a datos
    for row_idx in range(2, ws.max_row + 1):
        grupo_cell = ws.cell(row=row_idx, column=1)
        grupo_value = str(grupo_cell.value) if grupo_cell.value else ""
        
        # Detectar líneas de profesor
        if 'PROFESOR' in grupo_value:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = profesor_fill
                cell.font = profesor_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_thin
        else:
            # Aplicar formato normal y detectar salones inválidos
            tiene_invalido = False
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = str(cell.value) if cell.value else ""
                
                # Detectar salones inválidos
                if any(inv in cell_value for inv in ['AV1', 'AV2', 'AV4', 'AV5', 'E11']):
                    tiene_invalido = True
                    cell.fill = invalido_fill
                    cell.font = invalido_font
                
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border_thin
            
            # Si tiene salón inválido, resaltar toda la fila
            if tiene_invalido:
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.fill.start_color.rgb != invalido_fill.start_color.rgb:
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Ajustar anchos de columnas
    column_widths = {
        'A': 20,  # Grupo
        'B': 50,  # Materia
        'C': 15,  # Lunes
        'D': 15,  # Martes
        'E': 15,  # Miércoles
        'F': 15,  # Jueves
        'G': 15,  # Viernes
        'H': 10   # Class
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Ajustar altura de filas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20
    
    # Altura del encabezado
    ws.row_dimensions[1].height = 30
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar
    output_path = f"/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/Sistema-Salones-ISC/datos_estructurados/{nombre_salida}.xlsx"
    wb.save(output_path)
    
    print(f"   ✅ Excel formateado guardado: {output_path}")
    print(f"   📏 Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
    
    # Contar salones inválidos
    invalidos = 0
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(3, 8):  # Columnas de días
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = str(cell.value) if cell.value else ""
            if any(inv in cell_value for inv in ['AV1', 'AV2', 'AV4', 'AV5', 'E11']):
                invalidos += 1
    
    print(f"   ⚠️  Asignaciones inválidas detectadas: {invalidos}")
    
    return output_path

def main():
    """Función principal"""
    print("🎨 FORMATEADOR DE HORARIOS - CSV a Excel con Estilo")
    print("="*80)
    
    import os
    os.makedirs("/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/Sistema-Salones-ISC/datos_estructurados", exist_ok=True)
    
    # Formatear archivo original
    archivo_original = "/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/PROBLEMA SALONES ISC TEC/HorariosAgoDic2025.csv"
    path1 = aplicar_formato_horario(archivo_original, "Horario_Inicial_Formateado")
    
    # Formatear archivo optimizado
    archivo_optimizado = "/Users/lic.ing.jesusolvera/Documents/PROYECTOS PERSONALES/PROBLEMA SALONES ISC TEC/grouped_optimized_schedule.csv"
    path2 = aplicar_formato_horario(archivo_optimizado, "Horario_Optimizado_Formateado")
    
    print("\n" + "="*80)
    print("✅ ARCHIVOS FORMATEADOS EXITOSAMENTE")
    print("="*80)
    print("\n📁 Ubicación: Sistema-Salones-ISC/datos_estructurados/")
    print("\n📊 Archivos creados:")
    print("   1. Horario_Inicial_Formateado.xlsx")
    print("   2. Horario_Optimizado_Formateado.xlsx")
    
    print("\n🎨 Formato aplicado:")
    print("   ✅ Encabezados en azul con texto blanco")
    print("   ✅ Líneas de PROFESOR en naranja")
    print("   ✅ Salones inválidos (AV/E11) en rojo")
    print("   ✅ Columnas con ancho apropiado")
    print("   ✅ Bordes en todas las celdas")
    print("   ✅ Primera fila congelada")
    print("   ✅ Texto centrado y legible")
    
    print("\n🚀 Abre los archivos en Excel para ver el formato mejorado\n")

if __name__ == "__main__":
    main()
