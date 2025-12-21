#!/usr/bin/env python3
"""
Generador de Excel Comparativo de Salones - Agrupado por Materia
Para cada grupo/materia, muestra las 5 versiones juntas para comparación rápida
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def formatear_hora(bloque):
    """Convierte bloque horario a formato legible"""
    bloque = str(bloque).zfill(4)
    if len(bloque) == 4:
        hora_inicio_num = int(bloque[:2])
        hora_fin_num = int(bloque[2:])
        
        # Convertir a formato 12 horas
        if hora_inicio_num == 0:
            hora_inicio_str = "12:00AM"
        elif hora_inicio_num < 12:
            hora_inicio_str = f"{hora_inicio_num}:00AM"
        elif hora_inicio_num == 12:
            hora_inicio_str = "12:00PM"
        else:
            hora_inicio_str = f"{hora_inicio_num - 12}:00PM"
        
        if hora_fin_num == 0:
            hora_fin_str = "12:00AM"
        elif hora_fin_num < 12:
            hora_fin_str = f"{hora_fin_num}:00AM"
        elif hora_fin_num == 12:
            hora_fin_str = "12:00PM"
        else:
            hora_fin_str = f"{hora_fin_num - 12}:00PM"
        
        return f"{hora_inicio_str} - {hora_fin_str}"
    return bloque

def crear_fila_horario(grupo_df, metodo_nombre):
    """Crea una fila de horario para un método específico"""
    fila = {
        'Metodo': metodo_nombre,
        'Lunes': '',
        'Martes': '',
        'Miercoles': '',
        'Jueves': '',
        'Viernes': ''
    }
    
    for _, row in grupo_df.iterrows():
        dia = row['Dia']
        if dia == 'Sabado':
            continue
            
        bloque = row['Bloque_Horario']
        salon = row['Salon']
        
        hora_formateada = formatear_hora(bloque)
        celda = f"{hora_formateada} {salon}"
        
        if dia in fila:
            if fila[dia]:
                fila[dia] += f" {celda}"
            else:
                fila[dia] = celda
    
    return [fila['Metodo'], fila['Lunes'], fila['Martes'], fila['Miercoles'], fila['Jueves'], fila['Viernes']]

def main():
    print("\n" + "="*80)
    print("📊 GENERANDO EXCEL COMPARATIVO - AGRUPADO POR MATERIA")
    print("="*80)
    
    # Cargar todos los horarios
    print("\n📂 Cargando datos...")
    df_inicial = pd.read_csv("datos_estructurados/01_Horario_Inicial.csv")
    df_profesor = pd.read_csv("datos_estructurados/02_Horario_Optimizado_Profesor.csv")
    df_ml = pd.read_csv("datos_estructurados/03_Horario_Optimizado_ML.csv")
    df_genetico = pd.read_csv("datos_estructurados/04_Horario_Optimizado_Genetico.csv")
    df_greedy = pd.read_csv("datos_estructurados/04_Horario_Optimizado_Greedy.csv")
    
    print(f"   ✅ Todos los métodos cargados")
    
    # Crear Excel
    print("\n🔄 Creando comparativa agrupada...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativa Salones"
    
    # Encabezado
    encabezado = ['Método', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    ws.append(encabezado)
    
    # Agrupar por Grupo, Materia, Profesor
    grupos_materias = df_inicial.groupby(['Grupo', 'Materia', 'Profesor'])
    grupos_ordenados = sorted(grupos_materias.groups.keys(), key=lambda x: x[0])
    
    total_grupos = len(grupos_ordenados)
    print(f"   📋 Procesando {total_grupos} grupos...")
    
    for idx, (grupo, materia, profesor) in enumerate(grupos_ordenados, 1):
        if idx % 20 == 0:
            print(f"      Procesado {idx}/{total_grupos}...")
        
        # Título del grupo
        titulo = f"{grupo} - {materia} - {profesor}"
        ws.append([titulo, '', '', '', '', ''])
        
        # Obtener datos de este grupo para cada método
        grupo_inicial = df_inicial[(df_inicial['Grupo'] == grupo) & 
                                   (df_inicial['Materia'] == materia) & 
                                   (df_inicial['Profesor'] == profesor)]
        
        grupo_profesor = df_profesor[(df_profesor['Grupo'] == grupo) & 
                                     (df_profesor['Materia'] == materia) & 
                                     (df_profesor['Profesor'] == profesor)]
        
        grupo_ml = df_ml[(df_ml['Grupo'] == grupo) & 
                        (df_ml['Materia'] == materia) & 
                        (df_ml['Profesor'] == profesor)]
        
        grupo_genetico = df_genetico[(df_genetico['Grupo'] == grupo) & 
                                     (df_genetico['Materia'] == materia) & 
                                     (df_genetico['Profesor'] == profesor)]
        
        grupo_greedy = df_greedy[(df_greedy['Grupo'] == grupo) & 
                                 (df_greedy['Materia'] == materia) & 
                                 (df_greedy['Profesor'] == profesor)]
        
        # Agregar las 5 filas de comparación
        ws.append(crear_fila_horario(grupo_inicial, "🔵 Original"))
        ws.append(crear_fila_horario(grupo_profesor, "👨‍🏫 Profesor"))
        ws.append(crear_fila_horario(grupo_ml, "🤖 ML"))
        ws.append(crear_fila_horario(grupo_genetico, "🧬 Genético"))
        ws.append(crear_fila_horario(grupo_greedy, "⚡ Greedy"))
        
        # Fila vacía de separación
        ws.append(['', '', '', '', '', ''])
    
    print("   ✅ Todos los grupos procesados")
    
    # Aplicar formato
    print("\n🎨 Aplicando formato...")
    
    # Estilos
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    titulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    titulo_font = Font(bold=True, color="FFFFFF", size=12)
    
    original_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    profesor_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ml_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    genetico_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    greedy_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Aplicar formato a encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    
    # Aplicar formato a datos
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        primera_celda = str(row[0].value) if row[0].value else ''
        
        # Detectar tipo de fila
        es_titulo = primera_celda and not any(emoji in primera_celda for emoji in ['🔵', '👨‍🏫', '🤖', '🧬', '⚡'])
        es_separador = not primera_celda.strip()
        
        if es_titulo:
            # Título de grupo/materia
            for cell in row:
                cell.fill = titulo_fill
                cell.font = titulo_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border_thin
            ws.merge_cells(f'A{row_idx}:F{row_idx}')
        elif es_separador:
            # Separador
            for cell in row:
                cell.border = border_thin
        else:
            # Fila de método - aplicar color según método
            if '🔵' in primera_celda:
                fill = original_fill
            elif '👨‍🏫' in primera_celda:
                fill = profesor_fill
            elif '🤖' in primera_celda:
                fill = ml_fill
            elif '🧬' in primera_celda:
                fill = genetico_fill
            elif '⚡' in primera_celda:
                fill = greedy_fill
            else:
                fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for col_idx, cell in enumerate(row, 1):
                cell.fill = fill
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.font = Font(bold=True, size=10)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                cell.border = border_thin
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 25  # Método
    ws.column_dimensions['B'].width = 25  # Lunes
    ws.column_dimensions['C'].width = 25  # Martes
    ws.column_dimensions['D'].width = 25  # Miércoles
    ws.column_dimensions['E'].width = 25  # Jueves
    ws.column_dimensions['F'].width = 25  # Viernes
    
    # Ajustar altura de filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        primera_celda = str(row[0].value) if row[0].value else ''
        es_titulo = primera_celda and not any(emoji in primera_celda for emoji in ['🔵', '👨‍🏫', '🤖', '🧬', '⚡'])
        
        if es_titulo:
            ws.row_dimensions[row[0].row].height = 25
        else:
            ws.row_dimensions[row[0].row].height = 40
    
    ws.row_dimensions[1].height = 25
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar
    output_dir = Path("comparativas/excel_comparativo")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / "Comparativa_Salones_Todos_Metodos.xlsx"
    
    wb.save(output_file)
    
    print(f"   ✅ Excel guardado: {output_file}")
    
    # Estadísticas
    print("\n" + "="*80)
    print("📊 ESTADÍSTICAS")
    print("="*80)
    
    print(f"\n📋 Estructura del Excel:")
    print(f"   • {total_grupos} grupos/materias comparados")
    print(f"   • Cada grupo muestra 5 métodos juntos")
    print(f"   • Total de filas: {ws.max_row}")
    print(f"   • Colores diferentes para cada método")
    
    print("\n🎨 Código de colores:")
    print("   🔵 Original - Gris")
    print("   👨‍🏫 Profesor - Amarillo claro")
    print("   🤖 ML - Verde claro")
    print("   🧬 Genético - Morado claro")
    print("   ⚡ Greedy - Naranja claro")
    
    print("\n" + "="*80)
    print("✅ EXCEL COMPARATIVO GENERADO EXITOSAMENTE")
    print("="*80)
    print(f"\n📁 Ubicación: {output_file}")
    print("\n💡 Cómo usar:")
    print("   • Cada grupo/materia muestra los 5 métodos juntos")
    print("   • Compara rápidamente los salones asignados")
    print("   • Colores ayudan a identificar cada método")
    print("   • Formato compacto y fácil de leer\n")

if __name__ == "__main__":
    main()
